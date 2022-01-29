#!/usr/bin/env python

### Script to process metadata file - 03/08/2021 
# Python version 3.6
# Takes as input 'Tm1b+oim Tendon Diameter +sampleInfo_270721.2.xlsx' spreadsheet
# Outputs a processed metadata file for all samples, containing data from all sheets
# Dates in input spreadsheet need to be seperated by '.' instead of '/' so they can be matched to the sheet names
# Run in the directory the file is in
# Contact Emily Johnson at ejohn16@liv.ac.uk if you're having trouble with the script 

## Load packages

import pandas as pd
import os
import sys 
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
import re

# If any packages aren't installed use package manager (preferably anaconda) to install, e.g.
# conda install -c anaconda openpyxl

## Read in data and assign different sheets to a list variable that can be accessed/iterated over 

# Get current working directory 
dir = os.getcwd()
files = os.listdir(dir)

# Read in sheet WITHOUT formulas (important to include 'data_only=True')
wb = openpyxl.load_workbook("Tm1b+oim Tendon Diameter +sampleInfo_270721.2.xlsx", data_only=True)

# Assign all sheet-names to list 
sheets = wb.get_sheet_names()


## Convert sheet 1 containing the metadata into a dataframe
# Also makes sure all values in the dataframe are strings and removes empty rows
metadata = sheets[0]
metadata = wb[metadata]

values = metadata.values
df = pd.DataFrame(values)
df = df.dropna()
df = df.drop([3, 6], axis=1)


## Create summary array for data to be appended to 
summary = np.array(['Date', 'Sample_ID', '.', 'Sex', 'Age', 'Genotype', 'Replicate', 'Average_diameter', 'Circumference', 'Circumference_true', 'Date_ID'])
summary = np.reshape(summary, (1,11))


## Process the different sheets
# For each sheet same ID A,B and C will have the data extracted and reformatted using pyxl in combination with standard python data science tools
for sheet in sheets[1:31]:
    print(sheet)
    current_sheet = wb[sheet]

    ## ID A

    # Extract data from sheet

    A_rep_number = []
    for row in current_sheet.iter_rows(min_row=1,
                            max_row=1,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            A_rep_number.append(row)

    A_average_diameter = []
    for row in current_sheet.iter_rows(min_row=6,
                            max_row=6,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            A_average_diameter.append(row)

    A_circumference = []
    for row in current_sheet.iter_rows(min_row=10,
                            max_row=10,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            A_circumference.append(row)

    A_circumference_true = []
    for row in current_sheet.iter_rows(min_row=11,
                            max_row=11,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            A_circumference_true.append(row)

    # Create variables to search the metadata dataframe for the current sample 
    
    sam_id = str('A')
    date = sheet

    # Find the row in the metadata dataframe containing the sample ID and date specified in the above variables 
    # Convert this row to a numpy array, then tile it so it can be joined to the rest of the data

    contain_values = df[df[0].str.contains(date) & df[1].str.contains(sam_id)].to_numpy()
    repetitions = len(np.transpose(A_rep_number))
    repeats_array = np.tile(contain_values, (repetitions, 1))

    # Create a date ID column that can be used to index files from their file name in other scripts

    date_split = date.split('.')
    year = date_split[2]
    year = year[2:4]
    date_ID = str(year + date_split[1] + date_split[0])

    date_IDs = []
    for i in range(len(np.transpose(A_rep_number))):
        date_IDs.append(date_ID) 
    
    # Join it all together and append to summary

    A = np.concatenate((repeats_array, np.transpose(A_rep_number), np.transpose(A_average_diameter), np.transpose(A_circumference), np.transpose(A_circumference_true), np.reshape(date_IDs, (len(date_IDs),1))), axis=1)
    summary = np.concatenate((summary, A), axis=0)

    ## ID B

    # Extract data from sheet

    B_rep_number = []
    for row in current_sheet.iter_rows(min_row=16,
                            max_row=16,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            B_rep_number.append(row)


    B_average_diameter = []
    for row in current_sheet.iter_rows(min_row=21,
                            max_row=21,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            B_average_diameter.append(row)


    B_circumference = []
    for row in current_sheet.iter_rows(min_row=25,
                            max_row=25,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            B_circumference.append(row)

    B_circumference_true = []
    for row in current_sheet.iter_rows(min_row=26,
                            max_row=26,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            B_circumference_true.append(row)

    # Create variables to search the metadata dataframe for the current sample 

    sam_id = str('B')
    date = sheet

    # Find the row in the metadata dataframe containing the sample ID and date specified in the above variables 
    # Convert this row to a numpy array, then tile it so it can be joined to the rest of the data

    contain_values = df[df[0].str.contains(date) & df[1].str.contains(sam_id)].to_numpy()
    repetitions = len(np.transpose(B_rep_number))
    repeats_array = np.tile(contain_values, (repetitions, 1))

    # Date ID doesn't need to be recreated, its the same for all samples on the same sheet
    # Join it all together and append to summary

    B = np.concatenate((repeats_array, np.transpose(B_rep_number), np.transpose(B_average_diameter), np.transpose(B_circumference), np.transpose(B_circumference_true), np.reshape(date_IDs, (len(date_IDs),1))), axis=1)
    summary = np.concatenate((summary, B), axis=0)

    ## ID C

    # Extract data from sheet

    C_rep_number = []
    for row in current_sheet.iter_rows(min_row=31,
                            max_row=31,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            C_rep_number.append(row)

    C_average_diameter = []
    for row in current_sheet.iter_rows(min_row=36,
                            max_row=36,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            C_average_diameter.append(row)

    
    C_circumference = []
    for row in current_sheet.iter_rows(min_row=40,
                            max_row=40,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            C_circumference.append(row)

    C_circumference_true = []
    for row in current_sheet.iter_rows(min_row=41,
                            max_row=41,
                            min_col=3,
                            max_col=13,
                            values_only=True):
                            C_circumference_true.append(row)

    # Create variables to search the metadata dataframe for the current sample 

    sam_id = str('C')
    date = sheet

    # Find the row in the metadata dataframe containing the sample ID and date specified in the above variables 
    # Convert this row to a numpy array, then tile it so it can be joined to the rest of the data

    contain_values = df[df[0].str.contains(date) & df[1].str.contains(sam_id)].to_numpy()
    repetitions = len(np.transpose(C_rep_number))
    repeats_array = np.tile(contain_values, (repetitions, 1))

    # Date ID doesn't need to be recreated, its the same for all samples on the same sheet
    # Join it all together and append to summary

    C = np.concatenate((repeats_array, np.transpose(C_rep_number), np.transpose(C_average_diameter), np.transpose(C_circumference), np.transpose(C_circumference_true), np.reshape(date_IDs, (len(date_IDs),1))), axis=1)
    summary = np.concatenate((summary, C), axis=0)

    ## ID D

    # Check if sample D is present in sheet, then extract data and append to summary as previously

    if current_sheet["A46"].value == 'D':
        print("Sample D present")
        D_rep_number = []
        for row in current_sheet.iter_rows(min_row=46,
                                max_row=46,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                D_rep_number.append(row)


        D_average_diameter = []
        for row in current_sheet.iter_rows(min_row=51,
                                max_row=51,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                D_average_diameter.append(row)

        D_circumference = []
        for row in current_sheet.iter_rows(min_row=55,
                                max_row=55,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                D_circumference.append(row)

        D_circumference_true = []
        for row in current_sheet.iter_rows(min_row=56,
                                max_row=56,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                D_circumference_true.append(row)

        sam_id = str('D')
        date = sheet

        contain_values = df[df[0].str.contains(date) & df[1].str.contains(sam_id)].to_numpy()
        repetitions = len(np.transpose(D_rep_number))
        repeats_array = np.tile(contain_values, (repetitions, 1))

        D = np.concatenate((repeats_array, np.transpose(D_rep_number), np.transpose(D_average_diameter), np.transpose(D_circumference), np.transpose(D_circumference_true), np.reshape(date_IDs, (len(date_IDs),1))), axis=1)
        summary = np.concatenate((summary, D), axis=0)


    else:
        print("No sample ID D!")

    ## ID E

    # Check if sample E is present in sheet, then extract data and append to summary as previously

    if current_sheet["A61"].value == 'E':
        print("Sample E present")
        E_rep_number = []
        for row in current_sheet.iter_rows(min_row=61,
                                max_row=61,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                E_rep_number.append(row)


        E_average_diameter = []
        for row in current_sheet.iter_rows(min_row=66,
                                max_row=66,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                E_average_diameter.append(row)

        
        E_circumference = []
        for row in current_sheet.iter_rows(min_row=70,
                                max_row=70,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                E_circumference.append(row)

        E_circumference_true = []
        for row in current_sheet.iter_rows(min_row=71,
                                max_row=71,
                                min_col=3,
                                max_col=13,
                                values_only=True):
                                E_circumference_true.append(row)


        sam_id = str('E')
        date = sheet

        contain_values = df[df[0].str.contains(date) & df[1].str.contains(sam_id)].to_numpy()
        repetitions = len(np.transpose(E_rep_number))
        repeats_array = np.tile(contain_values, (repetitions, 1))

        E = np.concatenate((repeats_array, np.transpose(E_rep_number), np.transpose(E_average_diameter), np.transpose(E_circumference), np.transpose(E_circumference_true), np.reshape(date_IDs, (len(date_IDs),1))), axis=1)
        summary = np.concatenate((summary, E), axis=0)

    else:
        print("No sample ID E!")

## Convert the full summary array into a pandas dataframe

df = pd.DataFrame(summary)

# Process dataframe to have row headings
new_header = df.iloc[0]
df = df[1:]
df.columns = new_header

# Save the processed dataframe to a .csv file 
df.to_csv('tendon_data_formatted.csv', index=False)
